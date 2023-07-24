import os
import tkinter as tk
from tkinter import ttk
import openpyxl
from datetime import datetime, timedelta
from tkinter import messagebox
import time

class Application(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title("Job Application Tracker")
        self.geometry("550x300")  # Slightly bigger window size
        self.configure(bg="#F5F5F5")

        self.applications_sent_counter = 0
        self.applications_rejected_counter = 0
        self.timer_paused = False
        self.start_time = time.time()
        self.paused_time = 0
        self.elapsed_time = 0
        self.filename = self.get_file_name()

        self.style = ttk.Style()
        self.style.configure("TButton", font=("Arial", 12), padding=6)

        self.sent_label = ttk.Label(self, text="Applications Sent:", font=("Arial", 14), padding=(10, 10))
        self.sent_label.grid(row=0, column=0)

        self.sent_value_label = ttk.Label(self, text="0", font=("Arial", 14), padding=(10, 10))
        self.sent_value_label.grid(row=0, column=1)

        self.sent_increment_button = ttk.Button(self, text="+", command=self.increment_sent)
        self.sent_increment_button.grid(row=0, column=2, padx=5)

        self.sent_decrement_button = ttk.Button(self, text="-", command=self.decrement_sent)
        self.sent_decrement_button.grid(row=0, column=3, padx=5)

        self.rejected_label = ttk.Label(self, text="Applications Rejected:", font=("Arial", 14), padding=(10, 10))
        self.rejected_label.grid(row=1, column=0)

        self.rejected_value_label = ttk.Label(self, text="0", font=("Arial", 14), padding=(10, 10))
        self.rejected_value_label.grid(row=1, column=1)

        self.rejected_increment_button = ttk.Button(self, text="+", command=self.increment_rejected)
        self.rejected_increment_button.grid(row=1, column=2, padx=5)

        self.rejected_decrement_button = ttk.Button(self, text="-", command=self.decrement_rejected)
        self.rejected_decrement_button.grid(row=1, column=3, padx=5)

        self.save_and_exit_button = ttk.Button(self, text="Save and Exit", command=self.save_and_exit)
        self.save_and_exit_button.grid(row=2, column=0, columnspan=4, pady=10)

        self.pause_button = ttk.Button(self, text="Pause", command=self.toggle_pause)
        self.pause_button.grid(row=3, column=0, columnspan=4, pady=5)

        self.timer_label = ttk.Label(self, text="Time Elapsed: 00:00:00", font=("Arial", 12), padding=(0, 5))
        self.timer_label.grid(row=4, column=0, columnspan=4)

        # Bind the self-close button to the on_close method
        self.protocol("WM_DELETE_WINDOW", self.on_close)

        # Start the timer
        self.update_timer_label()

    def get_file_name(self):
        current_datetime = datetime.now()
        date_format = current_datetime.strftime("%Y-%m-%d")
        filename = f"JobApp - {date_format}.xlsx"
        return filename

    def update_timer_label(self):
        if not self.timer_paused:
            self.elapsed_time = int(time.time() - self.start_time - self.paused_time)
        formatted_time = str(timedelta(seconds=self.elapsed_time))
        self.timer_label.config(text=f"Time Elapsed: {formatted_time}")
        self.after(1000, self.update_timer_label)

    def toggle_pause(self):
        if not self.timer_paused:
            # Pause the timer
            self.timer_paused = True
            self.pause_button.config(text="Resume")

            # Record the current time as the pause start time
            self.pause_start_time = time.time()
        else:
            # Resume the timer
            self.timer_paused = False
            self.pause_button.config(text="Pause")

            # Accumulate the time spent in paused state
            self.paused_time += time.time() - self.pause_start_time

    def increment_sent(self):
        self.applications_sent_counter += 1
        self.update_counters()

    def decrement_sent(self):
        if self.applications_sent_counter > 0:
            self.applications_sent_counter -= 1
            self.update_counters()

    def increment_rejected(self):
        self.applications_rejected_counter += 1
        self.update_counters()

    def decrement_rejected(self):
        if self.applications_rejected_counter > 0:
            self.applications_rejected_counter -= 1
            self.update_counters()

    def update_counters(self):
        self.sent_value_label.config(text=str(self.applications_sent_counter))
        self.rejected_value_label.config(text=str(self.applications_rejected_counter))

    def save_to_spreadsheet(self):
        current_datetime = datetime.now()
        date_format = current_datetime.strftime("%Y-%m-%d")
        if date_format not in self.filename:
            # Date changed, update the filename
            self.filename = self.get_file_name()

        if os.path.isfile(self.filename):
            # File already exists, load it and update data
            workbook = openpyxl.load_workbook(self.filename)
            sheet = workbook.active

            # Find the row containing the current date's data
            for row in sheet.iter_rows(values_only=True):
                if row[0] == self.applications_sent_counter and row[1] == self.applications_rejected_counter:
                    # Data for the current date already exists, update the time taken and save the file
                    sheet.cell(row=sheet.max_row, column=3, value=str(timedelta(seconds=self.elapsed_time)))
                    workbook.save(self.filename)
                    messagebox.showinfo("Success", f"Excel sheet updated for {date_format}.")
                    return

        # Data for the current date doesn't exist, append it to the end
        workbook = openpyxl.load_workbook(self.filename)
        sheet = workbook.active

        # Write the data for the current date
        sheet.append([self.applications_sent_counter, self.applications_rejected_counter, str(timedelta(seconds=self.elapsed_time))])

        # Save the data to a file with the correct file extension
        workbook.save(self.filename)
        messagebox.showinfo("Success", f"Excel sheet saved as: {self.filename}")

    def on_close(self):
        # Call the save_to_spreadsheet method before destroying the application
        self.save_to_spreadsheet()
        self.destroy()

    def save_and_exit(self):
        # Call the save_to_spreadsheet method before closing the application
        self.save_to_spreadsheet()
        self.destroy()

if __name__ == "__main__":
    app = Application()
    app.mainloop()